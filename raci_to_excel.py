#!/usr/bin/env python3
"""
RACI Matrix to Excel Converter
Converts RACI-Matrix.md to a multi-sheet Excel workbook with professional formatting.
"""

import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class RACIExcelConverter:
    """Converts RACI markdown tables to formatted Excel workbook."""

    def __init__(self, md_file_path, output_path):
        self.md_file_path = Path(md_file_path)
        self.output_path = Path(output_path)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

        # Color scheme for RACI values
        self.colors = {
            'R': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),  # Blue
            'A': PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid'),  # Orange
            'C': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # Yellow
            'I': PatternFill(start_color='A5A5A5', end_color='A5A5A5', fill_type='solid'),  # Gray
            'R/A': PatternFill(start_color='C65911', end_color='C65911', fill_type='solid'),  # Dark Orange
            'header': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),  # Dark Blue
        }

        # Font styles
        self.header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.cell_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        self.normal_font = Font(name='Calibri', size=10)

        # Border style
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    def parse_markdown_table(self, table_text):
        """Parse a markdown table into a list of lists."""
        lines = table_text.strip().split('\n')
        # Remove separator line (the one with |---|---|)
        lines = [line for line in lines if not re.match(r'^\|[\s\-:|]+\|$', line)]

        rows = []
        for line in lines:
            # Split by | and clean up
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            rows.append(cells)

        return rows

    def extract_tables_from_markdown(self):
        """Extract all tables from the markdown file organized by section."""
        with open(self.md_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        sections = {}
        current_section = None
        current_subsection = None

        # Split by major sections
        lines = content.split('\n')
        i = 0
        while i < len(lines):
            line = lines[i]

            # Detect section headers
            if line.startswith('## '):
                current_section = line[3:].strip()
                sections[current_section] = {'subsections': {}, 'tables': []}
                current_subsection = None
            elif line.startswith('### ') and current_section:
                current_subsection = line[4:].strip()
                sections[current_section]['subsections'][current_subsection] = []
            elif line.startswith('|') and current_section:
                # Found a table, extract it
                table_lines = [line]
                i += 1
                while i < len(lines) and lines[i].startswith('|'):
                    table_lines.append(lines[i])
                    i += 1

                table_text = '\n'.join(table_lines)
                table_data = self.parse_markdown_table(table_text)

                if current_subsection:
                    sections[current_section]['subsections'][current_subsection].append(table_data)
                else:
                    sections[current_section]['tables'].append(table_data)

                i -= 1  # Back up one since we'll increment at the end

            i += 1

        return sections

    def apply_raci_formatting(self, ws, row, col, value):
        """Apply color formatting based on RACI value."""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = self.cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border

        # Determine which color to apply
        value_upper = str(value).upper().strip()
        if value_upper in self.colors:
            cell.fill = self.colors[value_upper]
        elif 'R/A' in value_upper or 'A/R' in value_upper:
            cell.fill = self.colors['R/A']
        elif 'R' in value_upper and 'A' in value_upper:
            cell.fill = self.colors['R/A']
        elif 'R' in value_upper:
            cell.fill = self.colors['R']
        elif 'A' in value_upper:
            cell.fill = self.colors['A']
        elif 'C' in value_upper:
            cell.fill = self.colors['C']
        elif 'I' in value_upper:
            cell.fill = self.colors['I']

    def write_table_to_sheet(self, ws, table_data, start_row=1):
        """Write a table to a worksheet with formatting."""
        if not table_data:
            return start_row

        # Write headers
        for col, header in enumerate(table_data[0], start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.colors['header']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Write data rows
        for row_idx, row_data in enumerate(table_data[1:], start=start_row + 1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                if col_idx == 1:  # First column (Activity)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.font = self.normal_font
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = self.border
                else:  # RACI columns
                    self.apply_raci_formatting(ws, row_idx, col_idx, cell_value)

        # Auto-adjust column widths
        for col in range(1, len(table_data[0]) + 1):
            if col == 1:
                ws.column_dimensions[get_column_letter(col)].width = 50
            else:
                ws.column_dimensions[get_column_letter(col)].width = 15

        # Set row heights
        ws.row_dimensions[start_row].height = 30
        for row in range(start_row + 1, start_row + len(table_data)):
            ws.row_dimensions[row].height = 25

        return start_row + len(table_data) + 2  # Return next available row

    def create_overview_sheet(self):
        """Create the overview sheet with role definitions and RACI legend."""
        ws = self.wb.create_sheet("Overview")

        # Title
        ws['A1'] = 'RACI Matrix for AI/ML & Automation Initiatives'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5496')
        ws.merge_cells('A1:F1')

        ws['A2'] = 'Bank ABC - Innovation & Digitization Department'
        ws['A2'].font = Font(name='Calibri', size=12, color='2F5496')
        ws.merge_cells('A2:F2')

        # RACI Legend
        ws['A4'] = 'RACI Legend'
        ws['A4'].font = Font(name='Calibri', size=14, bold=True)

        legend_data = [
            ['Code', 'Meaning', 'Description'],
            ['R', 'Responsible', 'Person who does the work to complete the task'],
            ['A', 'Accountable', 'Person ultimately answerable for the completion and approval'],
            ['C', 'Consulted', 'Person whose input is sought (two-way communication)'],
            ['I', 'Informed', 'Person who is kept up-to-date on progress (one-way communication)'],
            ['R/A', 'Responsible & Accountable', 'Person who both does the work and is accountable']
        ]

        self.write_table_to_sheet(ws, legend_data, start_row=5)

        # Role abbreviations
        ws['A12'] = 'Role Abbreviations'
        ws['A12'].font = Font(name='Calibri', size=14, bold=True)

        abbrev_data = [
            ['Abbreviation', 'Full Role', 'Description'],
            ['BizSpon', 'Business Sponsor', 'Senior business leader sponsoring the initiative'],
            ['BizOwn', 'Business Owner', 'Day-to-day business point of contact'],
            ['DigProdHead', 'Head of Digital Products', 'Backlog management, requirements, UAT coordination'],
            ['AILead', 'AI Engineering Lead', 'Head of AI Engineering team'],
            ['BA', 'Business Analyst', 'Analyzes requirements, documents processes'],
            ['DS', 'Data Scientist', 'Develops AI/ML models'],
            ['MLE', 'ML Engineer', 'Implements ML pipelines, handles MLOps'],
            ['AutoDev', 'Automation Developer', 'Develops RPA bots, automation solutions'],
            ['SolArch', 'Solution Architect', 'Defines solution architecture'],
            ['DataEng', 'Data Engineer', 'Builds data pipelines'],
            ['ModelVal', 'Model Validator', 'Performs independent validation'],
            ['QA', 'Quality Assurance', 'Tests and validates solutions'],
            ['ItOps', 'IT Operations', 'Manages production infrastructure'],
            ['ItSec', 'IT Security', 'Reviews security requirements'],
            ['ItInfra', 'IT Infrastructure', 'Manages infrastructure'],
            ['Comply', 'Compliance', 'Ensures regulatory compliance'],
            ['Risk', 'Risk Management', 'Assesses and manages risks'],
            ['EntArch', 'Enterprise Architecture', 'Oversees enterprise architecture'],
            ['DataGov', 'Data Governance', 'Manages data governance'],
            ['SteerCom', 'Steering Committee', 'Senior governance body for Tier 1 initiatives']
        ]

        self.write_table_to_sheet(ws, abbrev_data, start_row=13)

    def create_sheet_structure(self):
        """Create the multi-sheet structure with all RACI tables."""
        sections = self.extract_tables_from_markdown()

        # Define sheet mapping
        sheet_mapping = {
            '1. Overview': 'overview',  # Will be created separately
            '2. Governance': ['4. Strategic & Operational Governance RACI'],
            '3. Gate 0 - Intake': ['5. Gate 0: Intake & Prioritization RACI'],
            '4. Gate 1 - Discovery': ['6. Gate 1: Discovery & Feasibility RACI'],
            '5. Gate 2 - Design': ['7. Gate 2: Design Phase RACI'],
            '6. Gate 3 - Build': ['8. Gate 3: Build Phase RACI'],
            '7. Gate 4 - Validation': ['9. Gate 4: Validation & Testing RACI'],
            '8. Gate 5 - Deployment': ['10. Gate 5: Deployment RACI'],
            '9. Post-Impl & Ops': ['11. Post-Implementation Review RACI', '12. Ongoing Operations & Maintenance RACI'],
            '10. Retraining & Retire': ['13. Retraining & Enhancement RACI', '14. Model/Bot Retirement RACI'],
            '11. Copilot Mgmt': ['15. Copilot Tools Management RACI'],
            '12. Hub Mgmt': ['16. AI & Automation Hub Management RACI'],
            '13. Variations': ['17. Initiative Type Variations', '18. Tier-Based Variations'],
            '14. Appendices': ['20. Appendices']
        }

        # Create overview sheet first
        self.create_overview_sheet()

        # Create other sheets
        for sheet_name, section_names in sheet_mapping.items():
            if sheet_name == '1. Overview':
                continue

            ws = self.wb.create_sheet(sheet_name)
            current_row = 1

            for section_name in section_names:
                # Find matching section
                for key, section_data in sections.items():
                    if section_name.lower() in key.lower():
                        # Write section title
                        ws.cell(row=current_row, column=1).value = key
                        ws.cell(row=current_row, column=1).font = Font(name='Calibri', size=14, bold=True, color='2F5496')
                        current_row += 2

                        # Write subsection tables
                        for subsection_name, tables in section_data['subsections'].items():
                            if tables:
                                # Write subsection name
                                ws.cell(row=current_row, column=1).value = subsection_name
                                ws.cell(row=current_row, column=1).font = Font(name='Calibri', size=12, bold=True, color='404040')
                                current_row += 1

                                # Write each table
                                for table in tables:
                                    current_row = self.write_table_to_sheet(ws, table, current_row)

                        # Write section-level tables (if any)
                        for table in section_data['tables']:
                            current_row = self.write_table_to_sheet(ws, table, current_row)

                        current_row += 2  # Add spacing between sections

    def convert(self):
        """Main conversion method."""
        print("Starting RACI to Excel conversion...")
        print(f"Reading from: {self.md_file_path}")

        self.create_sheet_structure()

        # Save the workbook
        self.wb.save(self.output_path)
        print(f"✓ Excel workbook created: {self.output_path}")
        print(f"  File size: {self.output_path.stat().st_size / 1024:.1f} KB")


def main():
    """Main function to convert RACI matrix to Excel."""
    source_file = Path('/Users/govind/AIEngineering/RACI-Matrix.md')
    output_file = Path('/Users/govind/AIEngineering/governance-docs-word/RACI-Matrix.xlsx')

    if not source_file.exists():
        print(f"Error: Source file not found: {source_file}")
        return

    converter = RACIExcelConverter(source_file, output_file)
    converter.convert()

    print("\nConversion complete!")
    print(f"Open the file: {output_file}")


if __name__ == '__main__':
    main()
